"""خروجی اکسل و CSV گزارش‌ها — راست‌به‌چپ و با عنوان‌های فارسی."""
from __future__ import annotations

import csv
import io
from datetime import date

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from app.core.jalali import jalali_long, jalali_str
from app.services.report_service import DaySummary, PeriodSummary

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(name="B Nazanin", size=12, bold=True, color="FFFFFF")
CELL_FONT = Font(name="B Nazanin", size=11)
TITLE_FONT = Font(name="B Nazanin", size=14, bold=True, color="1F4E79")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN = Side(style="thin", color="BFBFBF")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

STATUS_FILLS = {
    "حاضر": PatternFill("solid", fgColor="E2EFDA"),
    "غایب": PatternFill("solid", fgColor="FCE4E4"),
    "مرخصی": PatternFill("solid", fgColor="FFF2CC"),
    "مأموریت": PatternFill("solid", fgColor="DEEBF7"),
    "تعطیل رسمی": PatternFill("solid", fgColor="EDEDED"),
    "تعطیل هفتگی": PatternFill("solid", fgColor="F5F5F5"),
    "ناقص": PatternFill("solid", fgColor="FFE0B2"),
}


def _setup_sheet(ws: Worksheet, title: str, subtitle: str, columns: list[str]) -> int:
    """عنوان و سرستون‌ها را می‌نویسد و شماره ردیف شروع داده را برمی‌گرداند."""
    ws.sheet_view.rightToLeft = True
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, len(columns)))
    cell = ws.cell(row=1, column=1, value=title)
    cell.font = TITLE_FONT
    cell.alignment = CENTER
    ws.row_dimensions[1].height = 26

    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(1, len(columns)))
    sub = ws.cell(row=2, column=1, value=subtitle)
    sub.font = CELL_FONT
    sub.alignment = CENTER
    ws.row_dimensions[2].height = 20

    for idx, name in enumerate(columns, start=1):
        c = ws.cell(row=4, column=idx, value=name)
        c.fill = HEADER_FILL
        c.font = HEADER_FONT
        c.alignment = CENTER
        c.border = BORDER
        ws.column_dimensions[get_column_letter(idx)].width = max(12, min(28, len(name) + 6))
    ws.row_dimensions[4].height = 24
    ws.freeze_panes = "A5"
    return 5


def _write_row(ws: Worksheet, row: int, values: list, status_col: int | None = None) -> None:
    for idx, value in enumerate(values, start=1):
        c = ws.cell(row=row, column=idx, value=value)
        c.font = CELL_FONT
        c.alignment = CENTER
        c.border = BORDER
        if status_col is not None and idx == status_col:
            fill = STATUS_FILLS.get(str(value))
            if fill:
                c.fill = fill


def _autofilter(ws: Worksheet, columns: int, last_row: int) -> None:
    if last_row >= 5:
        ws.auto_filter.ref = f"A4:{get_column_letter(columns)}{last_row}"


# --------------------------------------------------------------------- گزارش دوره‌ای

SUMMARY_COLUMNS = [
    "ردیف", "کد پرسنلی", "نام و نام خانوادگی", "واحد", "سمت",
    "روز حاضر", "روز غایب", "مرخصی", "مأموریت", "ناقص",
    "تعطیل رسمی", "تعطیل هفتگی", "مرخصی ساعتی", "کارکرد", "موظفی",
    "تأخیر (دقیقه)", "دفعات تأخیر", "تعجیل (دقیقه)", "اضافه‌کاری", "درصد حضور",
]

DETAIL_COLUMNS = [
    "کد پرسنلی", "نام و نام خانوادگی", "تاریخ", "روز هفته", "وضعیت",
    "ورود", "خروج", "کارکرد", "موظفی",
    "تأخیر (دقیقه)", "تعجیل (دقیقه)", "اضافه‌کاری", "تعداد تردد", "توضیح",
]


def build_period_workbook(
    summaries: list[PeriodSummary],
    start: date,
    end: date,
    report_title: str,
    include_details: bool = True,
) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "خلاصه"
    subtitle = f"از {jalali_long(start)} تا {jalali_long(end)}"
    row = _setup_sheet(ws, report_title, subtitle, SUMMARY_COLUMNS)

    for i, ps in enumerate(summaries, start=1):
        _write_row(
            ws,
            row,
            [
                i,
                ps.personnel_code,
                ps.full_name,
                ps.department_name or "-",
                ps.position or "-",
                ps.present_days,
                ps.absent_days,
                ps.leave_days,
                ps.mission_days,
                ps.incomplete_days,
                ps.holiday_days,
                ps.weekend_days,
                _hhmm(ps.hourly_leave_minutes),
                _hhmm(ps.worked_minutes),
                _hhmm(ps.expected_minutes),
                ps.late_minutes,
                ps.late_count,
                ps.early_leave_minutes,
                _hhmm(ps.overtime_minutes),
                f"{ps.attendance_rate}%",
            ],
        )
        row += 1
    _autofilter(ws, len(SUMMARY_COLUMNS), row - 1)

    # ردیف جمع کل
    total_row = row + 1
    ws.cell(row=total_row, column=1, value="جمع کل").font = HEADER_FONT
    ws.cell(row=total_row, column=1).fill = HEADER_FILL
    ws.cell(row=total_row, column=1).alignment = CENTER
    totals = {
        6: sum(s.present_days for s in summaries),
        7: sum(s.absent_days for s in summaries),
        8: sum(s.leave_days for s in summaries),
        9: sum(s.mission_days for s in summaries),
        10: sum(s.incomplete_days for s in summaries),
        16: sum(s.late_minutes for s in summaries),
        17: sum(s.late_count for s in summaries),
        18: sum(s.early_leave_minutes for s in summaries),
    }
    for col, value in totals.items():
        c = ws.cell(row=total_row, column=col, value=value)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = CENTER
    c = ws.cell(row=total_row, column=14, value=_hhmm(sum(s.worked_minutes for s in summaries)))
    c.font, c.fill, c.alignment = HEADER_FONT, HEADER_FILL, CENTER
    c = ws.cell(row=total_row, column=19, value=_hhmm(sum(s.overtime_minutes for s in summaries)))
    c.font, c.fill, c.alignment = HEADER_FONT, HEADER_FILL, CENTER

    if include_details:
        ws2 = wb.create_sheet("ریز تردد روزانه")
        row2 = _setup_sheet(ws2, f"{report_title} — ریز روزانه", subtitle, DETAIL_COLUMNS)
        status_col = DETAIL_COLUMNS.index("وضعیت") + 1
        for ps in summaries:
            for d in ps.days:
                r = d.as_row()
                _write_row(
                    ws2,
                    row2,
                    [
                        ps.personnel_code,
                        ps.full_name,
                        r["jalali_date"],
                        r["weekday"],
                        r["status_fa"],
                        r["first_in"] or "-",
                        r["last_out"] or "-",
                        r["worked_hhmm"],
                        r["expected_hhmm"],
                        r["late_minutes"],
                        r["early_leave_minutes"],
                        r["overtime_hhmm"],
                        r["punch_count"],
                        r["note"],
                    ],
                    status_col=status_col,
                )
                row2 += 1
        _autofilter(ws2, len(DETAIL_COLUMNS), row2 - 1)

    return _save(wb)


# ------------------------------------------------------------------- گزارش یک روز

DAILY_COLUMNS = [
    "ردیف", "کد پرسنلی", "نام و نام خانوادگی", "واحد", "سمت", "شیفت",
    "وضعیت", "ساعت ورود", "ساعت خروج", "کارکرد",
    "تأخیر (دقیقه)", "تعجیل (دقیقه)", "اضافه‌کاری", "توضیح",
]


def build_daily_workbook(rows: list[dict], day: date, report_title: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "حضور و غیاب روزانه"
    row = _setup_sheet(ws, report_title, jalali_long(day), DAILY_COLUMNS)
    status_col = DAILY_COLUMNS.index("وضعیت") + 1
    for i, r in enumerate(rows, start=1):
        _write_row(
            ws,
            row,
            [
                i,
                r.get("personnel_code", ""),
                r.get("full_name", ""),
                r.get("department_name") or "-",
                r.get("position") or "-",
                r.get("shift_name") or "-",
                r.get("status_fa", ""),
                r.get("first_in") or "-",
                r.get("last_out") or "-",
                r.get("worked_hhmm", "0:00"),
                r.get("late_minutes", 0),
                r.get("early_leave_minutes", 0),
                r.get("overtime_hhmm", "0:00"),
                r.get("note", ""),
            ],
            status_col=status_col,
        )
        row += 1
    _autofilter(ws, len(DAILY_COLUMNS), row - 1)
    return _save(wb)


# ---------------------------------------------------------------------- ریز تردد

PUNCH_COLUMNS = [
    "ردیف", "کد پرسنلی", "نام و نام خانوادگی", "واحد",
    "تاریخ", "ساعت", "نوع", "روش ثبت", "دستگاه", "اطمینان", "آفلاین", "توضیح",
]


def build_punches_workbook(rows: list[dict], start: date, end: date) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "ریز ترددها"
    subtitle = f"از {jalali_long(start)} تا {jalali_long(end)}"
    row = _setup_sheet(ws, "گزارش ریز ترددها", subtitle, PUNCH_COLUMNS)
    for i, r in enumerate(rows, start=1):
        _write_row(
            ws,
            row,
            [
                i,
                r.get("personnel_code", ""),
                r.get("employee_name", ""),
                r.get("department_name") or "-",
                r.get("jalali_date", ""),
                r.get("clock", ""),
                r.get("kind_fa", ""),
                r.get("method_fa", ""),
                r.get("device_name") or "-",
                round(r["confidence"], 3) if r.get("confidence") is not None else "-",
                "بله" if r.get("created_offline") else "خیر",
                r.get("note") or "",
            ],
        )
        row += 1
    _autofilter(ws, len(PUNCH_COLUMNS), row - 1)
    return _save(wb)


# ------------------------------------------------------------------- گزارش وظایف

TASK_COLUMNS = [
    "ردیف", "عنوان وظیفه", "پرسنل", "کد پرسنلی", "واحد",
    "وضعیت", "اولویت", "تکرار", "سررسید", "پیشرفت", "زمان برآوردی (دقیقه)", "توضیحات",
]


def build_tasks_workbook(rows: list[dict], title: str, subtitle: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "وظایف"
    row = _setup_sheet(ws, title, subtitle, TASK_COLUMNS)
    for i, r in enumerate(rows, start=1):
        _write_row(
            ws,
            row,
            [
                i,
                r.get("title", ""),
                r.get("employee_name") or "-",
                r.get("personnel_code") or "-",
                r.get("department_name") or "-",
                r.get("status_fa", ""),
                r.get("priority_fa", ""),
                r.get("recurrence_fa", ""),
                r.get("due_jalali_date") or "-",
                f"{r.get('progress', 0)}%",
                r.get("estimated_minutes") or "-",
                r.get("description") or "",
            ],
        )
        row += 1
    _autofilter(ws, len(TASK_COLUMNS), row - 1)
    return _save(wb)


# ------------------------------------------------------------------------- کمکی


def _hhmm(minutes: int | None) -> str:
    minutes = int(minutes or 0)
    return f"{minutes // 60}:{minutes % 60:02d}"


def _save(wb: Workbook) -> bytes:
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def to_csv(rows: list[dict], columns: list[tuple[str, str]]) -> bytes:
    """CSV با BOM تا اکسل فارسی درست باز کند. columns = [(کلید، عنوان فارسی)]"""
    buf = io.StringIO()
    writer = csv.writer(buf)
    writer.writerow([label for _, label in columns])
    for r in rows:
        writer.writerow([r.get(key, "") for key, _ in columns])
    return b"\xef\xbb\xbf" + buf.getvalue().encode("utf-8")


def filename_for(prefix: str, start: date, end: date | None = None) -> str:
    if end and end != start:
        return f"{prefix}_{jalali_str(start, '-')}_{jalali_str(end, '-')}.xlsx"
    return f"{prefix}_{jalali_str(start, '-')}.xlsx"
